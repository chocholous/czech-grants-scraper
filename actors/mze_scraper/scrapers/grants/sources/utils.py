"""
Document conversion utilities for grant sub-scrapers.

Converts PDF, XLSX, DOCX files to markdown format for LLM consumption.
"""

from pathlib import Path
from typing import Optional, Dict
import re

try:
    from apify import Actor
    ACTOR_AVAILABLE = True
except ImportError:
    ACTOR_AVAILABLE = False
    import logging

# Document download (using httpx for async support)
try:
    import httpx
    HTTPX_AVAILABLE = True
except ImportError:
    HTTPX_AVAILABLE = False
import requests

# PDF conversion
import pdfplumber

# Excel conversion
import pandas as pd
from openpyxl import load_workbook

# Word conversion
import mammoth
from markdownify import markdownify as md


# Well-named semantic constants
DOWNLOAD_CHUNK_SIZE_BYTES = 8192  # Standard chunk size for streaming downloads
DEFAULT_DOWNLOAD_TIMEOUT_SECONDS = 30  # Reasonable timeout for document downloads


if not ACTOR_AVAILABLE:
    logger = logging.getLogger(__name__)


def download_document(url: str, save_path: str, timeout: int = DEFAULT_DOWNLOAD_TIMEOUT_SECONDS) -> bool:
    """
    Download document from URL to local path.

    Args:
        url: Full URL to document
        save_path: Absolute path where file should be saved
        timeout: Request timeout in seconds

    Returns:
        True if download succeeded, False otherwise
    """
    try:
        save_path_obj = Path(save_path)
        save_path_obj.parent.mkdir(parents=True, exist_ok=True)

        if HTTPX_AVAILABLE:
            # Use httpx for better async compatibility
            with httpx.Client(timeout=timeout) as client:
                response = client.get(url)
                response.raise_for_status()
                
                with open(save_path, 'wb') as f:
                    f.write(response.content)
        else:
            # Fallback to requests
            response = requests.get(url, timeout=timeout, stream=True)
            response.raise_for_status()

        with open(save_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=DOWNLOAD_CHUNK_SIZE_BYTES):
                if chunk:
                    f.write(chunk)

        if ACTOR_AVAILABLE:
            Actor.log.info(f"Downloaded: {url} → {save_path}")
        else:
            logger.info(f"Downloaded: {url} → {save_path}")
        return True

    except Exception as e:
        if ACTOR_AVAILABLE:
            Actor.log.error(f"Failed to download {url}: {e}")
        else:
            logger.error(f"Failed to download {url}: {e}")
        return False


def convert_pdf_to_markdown(pdf_path: str) -> Optional[str]:
    """
    Convert PDF to markdown using pdfplumber.

    Extracts text and tables from all pages.

    Args:
        pdf_path: Path to PDF file

    Returns:
        Markdown string or None if conversion failed
    """
    try:
        markdown_parts = []

        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                # Extract text
                text = page.extract_text()
                if text:
                    markdown_parts.append(f"## Page {page_num}\n\n{text}\n")

                # Extract tables
                tables = page.extract_tables()
                if tables:
                    for table_num, table in enumerate(tables, start=1):
                        markdown_table = _table_to_markdown(table)
                        markdown_parts.append(f"### Table {table_num}\n\n{markdown_table}\n")

        result = "\n".join(markdown_parts)
        if ACTOR_AVAILABLE:
            Actor.log.info(f"Converted PDF to markdown: {pdf_path} ({len(result)} chars)")
        else:
            logger.info(f"Converted PDF to markdown: {pdf_path} ({len(result)} chars)")
        return result

    except Exception as e:
        if ACTOR_AVAILABLE:
            Actor.log.error(f"Failed to convert PDF {pdf_path}: {e}")
        else:
            logger.error(f"Failed to convert PDF {pdf_path}: {e}")
        return None


def convert_xlsx_to_markdown(xlsx_path: str) -> Optional[str]:
    """
    Convert Excel file to markdown.

    Extracts all sheets as markdown tables. For budget calculators,
    preserves formulas in a separate section.

    Args:
        xlsx_path: Path to XLSX file

    Returns:
        Markdown string or None if conversion failed
    """
    try:
        markdown_parts = []

        # Read with pandas for data
        excel_file = pd.ExcelFile(xlsx_path)

        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(xlsx_path, sheet_name=sheet_name)

            # Skip empty sheets
            if df.empty:
                continue

            markdown_parts.append(f"## Sheet: {sheet_name}\n")
            markdown_table = df.to_markdown(index=False)
            markdown_parts.append(f"{markdown_table}\n")

        # Extract formulas with openpyxl (if file contains formulas)
        try:
            wb = load_workbook(xlsx_path, data_only=False)
            formulas = _extract_formulas(wb)
            if formulas:
                markdown_parts.append("\n## Formulas\n")
                for cell_ref, formula in formulas.items():
                    markdown_parts.append(f"- `{cell_ref}`: `{formula}`")
        except Exception as e:
            if ACTOR_AVAILABLE:
                Actor.log.debug(f"Could not extract formulas from {xlsx_path}: {e}")
            else:
                logger.debug(f"Could not extract formulas from {xlsx_path}: {e}")

        result = "\n".join(markdown_parts)
        if ACTOR_AVAILABLE:
            Actor.log.info(f"Converted XLSX to markdown: {xlsx_path} ({len(result)} chars)")
        else:
            logger.info(f"Converted XLSX to markdown: {xlsx_path} ({len(result)} chars)")
        return result

    except Exception as e:
        if ACTOR_AVAILABLE:
            Actor.log.error(f"Failed to convert XLSX {xlsx_path}: {e}")
        else:
            logger.error(f"Failed to convert XLSX {xlsx_path}: {e}")
        return None


def convert_docx_to_markdown(docx_path: str) -> Optional[str]:
    """
    Convert DOCX to markdown using mammoth + markdownify pipeline.

    Args:
        docx_path: Path to DOCX file

    Returns:
        Markdown string or None if conversion failed
    """
    try:
        # Step 1: DOCX → HTML with mammoth
        with open(docx_path, "rb") as docx_file:
            result = mammoth.convert_to_html(docx_file)
            html = result.value

        # Step 2: HTML → Markdown with markdownify
        markdown = md(html, heading_style="ATX")

        # Clean up excessive newlines
        markdown = re.sub(r'\n{3,}', '\n\n', markdown)

        if ACTOR_AVAILABLE:
            Actor.log.info(f"Converted DOCX to markdown: {docx_path} ({len(markdown)} chars)")
        else:
            logger.info(f"Converted DOCX to markdown: {docx_path} ({len(markdown)} chars)")
        return markdown

    except Exception as e:
        if ACTOR_AVAILABLE:
            Actor.log.error(f"Failed to convert DOCX {docx_path}: {e}")
        else:
            logger.error(f"Failed to convert DOCX {docx_path}: {e}")
        return None


def convert_document_to_markdown(file_path: str) -> Optional[str]:
    """
    Auto-detect file type and convert to markdown.

    Args:
        file_path: Path to document (PDF, XLSX, DOCX)

    Returns:
        Markdown string or None if conversion failed or unsupported format
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    converters = {
        '.pdf': convert_pdf_to_markdown,
        '.xlsx': convert_xlsx_to_markdown,
        '.xlsm': convert_xlsx_to_markdown,  # Excel with macros
        '.docx': convert_docx_to_markdown,
    }

    converter = converters.get(suffix)
    if not converter:
        if ACTOR_AVAILABLE:
            Actor.log.warning(f"Unsupported file format: {suffix}")
        else:
            logger.warning(f"Unsupported file format: {suffix}")
        return None

    return converter(str(file_path))


# ===== Helper Functions =====

def _table_to_markdown(table: list) -> str:
    """
    Convert pdfplumber table (list of lists) to markdown table.

    Args:
        table: List of lists representing table rows

    Returns:
        Markdown table string
    """
    if not table or len(table) < 2:
        return ""

    # Use first row as headers
    headers = table[0]
    rows = table[1:]

    # Build markdown table
    markdown_lines = []

    # Header row
    header_row = "| " + " | ".join(str(cell or "") for cell in headers) + " |"
    markdown_lines.append(header_row)

    # Separator row
    separator = "| " + " | ".join("---" for _ in headers) + " |"
    markdown_lines.append(separator)

    # Data rows
    for row in rows:
        data_row = "| " + " | ".join(str(cell or "") for cell in row) + " |"
        markdown_lines.append(data_row)

    return "\n".join(markdown_lines)


def _extract_formulas(workbook) -> Dict[str, str]:
    """
    Extract all formulas from Excel workbook.

    Args:
        workbook: openpyxl Workbook object

    Returns:
        Dictionary mapping cell references to formulas (e.g., {"A1": "=SUM(B1:B10)"})
    """
    formulas = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                # Check if cell contains a formula
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_ref = f"{sheet_name}!{cell.coordinate}"
                    formulas[cell_ref] = cell.value

    return formulas
