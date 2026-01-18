"""
Excel parsing plugin using openpyxl and pandas.

Extracts data from Excel and CSV files.
"""

from pathlib import Path
from typing import Optional

import structlog

logger = structlog.get_logger(__name__)

# Check if dependencies are available
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


def read_excel_file(
    file_path: str,
    sheet_name: Optional[str] = None,
) -> Optional[list[list[str]]]:
    """
    Read Excel file into list of rows.

    Args:
        file_path: Path to Excel file
        sheet_name: Optional specific sheet name

    Returns:
        List of rows, where each row is a list of cell values
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl_not_available", hint="pip install openpyxl")
        return None

    path = Path(file_path)
    if not path.exists():
        logger.error("excel_not_found", path=file_path)
        return None

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name:
            sheet = wb[sheet_name]
        else:
            sheet = wb.active

        rows = []
        for row in sheet.iter_rows():
            row_data = [
                str(cell.value) if cell.value is not None else ""
                for cell in row
            ]
            rows.append(row_data)

        wb.close()

        logger.info(
            "excel_read",
            path=file_path,
            rows=len(rows),
        )

        return rows

    except Exception as e:
        logger.error("excel_read_failed", path=file_path, error=str(e))
        return None


def read_excel_as_dataframe(
    file_path: str,
    sheet_name: Optional[str] = None,
):
    """
    Read Excel file as pandas DataFrame.

    Args:
        file_path: Path to Excel file
        sheet_name: Optional specific sheet name

    Returns:
        pandas DataFrame or None
    """
    if not PANDAS_AVAILABLE:
        logger.error("pandas_not_available", hint="pip install pandas openpyxl")
        return None

    path = Path(file_path)
    if not path.exists():
        logger.error("excel_not_found", path=file_path)
        return None

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        logger.info(
            "excel_dataframe_read",
            path=file_path,
            rows=len(df),
            cols=len(df.columns),
        )

        return df

    except Exception as e:
        logger.error("excel_read_failed", path=file_path, error=str(e))
        return None


def read_csv_file(
    file_path: str,
    encoding: str = "utf-8",
    delimiter: str = ",",
) -> Optional[list[list[str]]]:
    """
    Read CSV file into list of rows.

    Args:
        file_path: Path to CSV file
        encoding: File encoding
        delimiter: Column delimiter

    Returns:
        List of rows
    """
    path = Path(file_path)
    if not path.exists():
        logger.error("csv_not_found", path=file_path)
        return None

    try:
        import csv

        rows = []
        with open(file_path, "r", encoding=encoding) as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row in reader:
                rows.append(row)

        logger.info(
            "csv_read",
            path=file_path,
            rows=len(rows),
        )

        return rows

    except Exception as e:
        logger.error("csv_read_failed", path=file_path, error=str(e))
        return None


def convert_excel_to_markdown(file_path: str) -> Optional[str]:
    """
    Convert Excel file to markdown format.

    Args:
        file_path: Path to Excel file

    Returns:
        Markdown string or None
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl_not_available")
        return None

    path = Path(file_path)
    if not path.exists():
        logger.error("excel_not_found", path=file_path)
        return None

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        md_parts = [f"# {path.name}\n"]

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            md_parts.append(f"\n## {sheet_name}\n")

            rows = []
            for row in sheet.iter_rows():
                row_data = [
                    str(cell.value) if cell.value is not None else ""
                    for cell in row
                ]
                rows.append(row_data)

            if rows:
                md_table = _rows_to_markdown_table(rows)
                md_parts.append(md_table)

        wb.close()

        markdown = "\n".join(md_parts)

        logger.info(
            "excel_converted_to_markdown",
            path=file_path,
            chars=len(markdown),
        )

        return markdown

    except Exception as e:
        logger.error("excel_conversion_failed", path=file_path, error=str(e))
        return None


def _rows_to_markdown_table(rows: list[list[str]]) -> str:
    """Convert rows to markdown table."""
    if not rows or not rows[0]:
        return ""

    num_cols = max(len(row) for row in rows)

    # Calculate column widths
    col_widths = [3] * num_cols
    for row in rows:
        for i, cell in enumerate(row):
            if i < num_cols:
                col_widths[i] = max(col_widths[i], len(str(cell)))

    lines = []

    # Header (first row)
    header = rows[0]
    header_cells = [
        str(cell).ljust(col_widths[i])
        for i, cell in enumerate(header[:num_cols])
    ]
    while len(header_cells) < num_cols:
        header_cells.append("".ljust(col_widths[len(header_cells)]))
    lines.append("| " + " | ".join(header_cells) + " |")

    # Separator
    separator = ["-" * w for w in col_widths]
    lines.append("| " + " | ".join(separator) + " |")

    # Data rows
    for row in rows[1:]:
        cells = [
            str(cell).ljust(col_widths[i])
            for i, cell in enumerate(row[:num_cols])
        ]
        while len(cells) < num_cols:
            cells.append("".ljust(col_widths[len(cells)]))
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


def get_excel_sheet_names(file_path: str) -> list[str]:
    """
    Get list of sheet names in Excel file.

    Args:
        file_path: Path to Excel file

    Returns:
        List of sheet names
    """
    if not OPENPYXL_AVAILABLE:
        return []

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []
